import json as _json

from flask import Blueprint, request, jsonify, session

from auth import require_module
from database import get_db

bp = Blueprint('finance', __name__)


def init_finance_db():
    migrations = [
        """CREATE TABLE IF NOT EXISTS finance_categories (
            id          SERIAL PRIMARY KEY,
            name        TEXT NOT NULL,
            type        TEXT NOT NULL DEFAULT 'expense',
            color       TEXT DEFAULT '#4a7bda',
            sort_order  INT DEFAULT 0,
            active      BOOLEAN DEFAULT TRUE,
            created_at  TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_records (
            id              SERIAL PRIMARY KEY,
            record_date     DATE NOT NULL,
            category_id     INT REFERENCES finance_categories(id) ON DELETE SET NULL,
            type            TEXT NOT NULL DEFAULT 'expense',
            title           TEXT NOT NULL,
            amount          NUMERIC(14,2) NOT NULL DEFAULT 0,
            tax_amount      NUMERIC(14,2) DEFAULT 0,
            vendor          TEXT DEFAULT '',
            invoice_no      TEXT DEFAULT '',
            note            TEXT DEFAULT '',
            document_id     INT,
            created_by      TEXT DEFAULT '',
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            updated_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_documents (
            id              SERIAL PRIMARY KEY,
            filename        TEXT NOT NULL,
            doc_type        TEXT DEFAULT '',
            ocr_raw         JSONB DEFAULT '{}',
            image_data      TEXT DEFAULT '',
            upload_date     DATE DEFAULT CURRENT_DATE,
            created_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_recurring (
            id              SERIAL PRIMARY KEY,
            title           TEXT NOT NULL,
            type            TEXT NOT NULL DEFAULT 'expense',
            category_id     INT REFERENCES finance_categories(id) ON DELETE SET NULL,
            amount          NUMERIC(14,2) NOT NULL DEFAULT 0,
            tax_amount      NUMERIC(14,2) DEFAULT 0,
            vendor          TEXT DEFAULT '',
            note            TEXT DEFAULT '',
            frequency       TEXT NOT NULL DEFAULT 'monthly',
            day_of_month    INT DEFAULT 1,
            start_date      DATE NOT NULL,
            end_date        DATE,
            last_generated  TEXT DEFAULT '',
            active          BOOLEAN DEFAULT TRUE,
            created_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS bank_statements (
            id                  SERIAL PRIMARY KEY,
            account_name        TEXT DEFAULT '',
            txn_date            DATE NOT NULL,
            amount              NUMERIC(14,2) NOT NULL,
            txn_type            TEXT DEFAULT 'debit',
            description         TEXT DEFAULT '',
            reconciled          BOOLEAN DEFAULT FALSE,
            matched_record_id   INT REFERENCES finance_records(id) ON DELETE SET NULL,
            import_batch        TEXT DEFAULT '',
            created_at          TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_payables (
            id              SERIAL PRIMARY KEY,
            payable_type    TEXT NOT NULL DEFAULT 'payable',
            title           TEXT NOT NULL,
            party_name      TEXT DEFAULT '',
            invoice_no      TEXT DEFAULT '',
            amount          NUMERIC(14,2) NOT NULL DEFAULT 0,
            due_date        DATE,
            status          TEXT NOT NULL DEFAULT 'open',
            paid_date       DATE,
            linked_record_id INT REFERENCES finance_records(id) ON DELETE SET NULL,
            note            TEXT DEFAULT '',
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            updated_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_budgets (
            id              SERIAL PRIMARY KEY,
            year            INT NOT NULL,
            month           INT NOT NULL,
            category_id     INT REFERENCES finance_categories(id) ON DELETE CASCADE,
            budget_amount   NUMERIC(14,2) NOT NULL DEFAULT 0,
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            updated_at      TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE(year, month, category_id)
        )""",
        "ALTER TABLE salary_records ADD COLUMN IF NOT EXISTS finance_synced BOOLEAN DEFAULT FALSE",
    ]
    for sql in migrations:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[finance_init] {str(e)[:80]}")

    defaults_income = [
        ('餐飲內用收入', 'income', '#2e9e6b', 1),
        ('外帶收入',     'income', '#0ea5e9', 2),
        ('外送收入',     'income', '#8b5cf6', 3),
        ('其他收入',     'income', '#c8a96e', 4),
    ]
    defaults_expense = [
        ('食材成本', 'expense', '#d64242', 10),
        ('薪資支出', 'expense', '#e07b2a', 11),
        ('租金',     'expense', '#8892a4', 12),
        ('水電費',   'expense', '#4a7bda', 13),
        ('設備維修', 'expense', '#e05c8a', 14),
        ('消耗品',   'expense', '#6366f1', 15),
        ('廣告行銷', 'expense', '#f59e0b', 16),
        ('其他支出', 'expense', '#64748b', 17),
    ]
    try:
        with get_db() as conn:
            cnt = conn.execute("SELECT COUNT(*) as c FROM finance_categories").fetchone()['c']
            if cnt == 0:
                for name, ftype, color, sort in (defaults_income + defaults_expense):
                    conn.execute(
                        "INSERT INTO finance_categories (name,type,color,sort_order) VALUES (%s,%s,%s,%s)",
                        (name, ftype, color, sort)
                    )
    except Exception as e:
        print(f"[finance_seed] {e}")


def init_finance_settings_db():
    migrations = [
        "ALTER TABLE finance_categories ADD COLUMN IF NOT EXISTS statement_section TEXT",
        """CREATE TABLE IF NOT EXISTS finance_settings (
            id            SERIAL PRIMARY KEY,
            setting_key   TEXT UNIQUE NOT NULL,
            setting_value TEXT DEFAULT ''
        )""",
    ]
    for sql in migrations:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[finance_settings_init] {str(e)[:80]}")

    section_defaults = {
        '餐飲內用收入': 'operating_revenue',
        '外帶收入':     'operating_revenue',
        '外送收入':     'operating_revenue',
        '其他收入':     'other_revenue',
        '食材成本':     'cogs',
        '薪資支出':     'operating_expense',
        '租金':         'operating_expense',
        '水電費':       'operating_expense',
        '設備維修':     'operating_expense',
        '消耗品':       'operating_expense',
        '廣告行銷':     'operating_expense',
        '其他支出':     'other_expense',
    }
    try:
        with get_db() as conn:
            for name, sec in section_defaults.items():
                conn.execute(
                    "UPDATE finance_categories SET statement_section=%s WHERE name=%s AND statement_section IS NULL",
                    (sec, name)
                )
            conn.execute("""
                UPDATE finance_categories
                SET statement_section = CASE WHEN type='income' THEN 'operating_revenue' ELSE 'operating_expense' END
                WHERE statement_section IS NULL
            """)
    except Exception as e:
        print(f"[finance_settings_seed] {e}")

    for k, v in [('company_name', ''), ('opening_cash', '0'), ('opening_equity', '0'),
                  ('company_tax_id', ''), ('company_address', '')]:
        try:
            with get_db() as conn:
                conn.execute(
                    "INSERT INTO finance_settings (setting_key, setting_value) VALUES (%s,%s) ON CONFLICT (setting_key) DO NOTHING",
                    (k, v)
                )
        except Exception as e:
            print(f"[finance_settings_default] {e}")


def init_insurance_db():
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS insurance_settings (
                    setting_key   TEXT PRIMARY KEY,
                    setting_value TEXT DEFAULT ''
                )
            """)
        for k, v in [('labor_insurance_no', ''), ('health_insurance_no', ''),
                     ('employer_name', ''), ('employer_id', '')]:
            with get_db() as conn:
                conn.execute(
                    "INSERT INTO insurance_settings (setting_key, setting_value) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                    (k, v))
    except Exception as e:
        print(f"[insurance_init] {e}")


# ── Row formatters ─────────────────────────────────────────────

def _finance_cat_row(r):
    if not r: return None
    d = dict(r)
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    return d


def _finance_rec_row(r):
    if not r: return None
    d = dict(r)
    if d.get('record_date'): d['record_date'] = str(d['record_date'])
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    if d.get('updated_at'): d['updated_at'] = d['updated_at'].isoformat()
    for f in ('amount', 'tax_amount'):
        if d.get(f) is not None: d[f] = float(d[f])
    return d


def _bank_row(r):
    if not r: return None
    d = dict(r)
    if d.get('txn_date'):   d['txn_date']   = str(d['txn_date'])
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    if d.get('amount') is not None: d['amount'] = float(d['amount'])
    return d


def _payable_row(r):
    if not r: return None
    d = dict(r)
    if d.get('due_date'):   d['due_date']   = str(d['due_date'])
    if d.get('paid_date'):  d['paid_date']  = str(d['paid_date'])
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    if d.get('updated_at'): d['updated_at'] = d['updated_at'].isoformat()
    if d.get('amount') is not None: d['amount'] = float(d['amount'])
    return d


def _get_finance_settings():
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT setting_key, setting_value FROM finance_settings").fetchall()
        return {r['setting_key']: r['setting_value'] for r in rows}
    except Exception:
        return {}


def _compute_statements(year, month):
    period = f"{year}-{str(month).zfill(2)}"
    with get_db() as conn:
        rows = conn.execute("""
            SELECT fr.type, fr.amount, fc.statement_section
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            WHERE TO_CHAR(fr.record_date,'YYYY-MM')=%s
        """, (period,)).fetchall()
        settings = _get_finance_settings()

    opening_cash   = float(settings.get('opening_cash',   '0') or '0')
    opening_equity = float(settings.get('opening_equity', '0') or '0')

    sections = {}
    for r in rows:
        sec = r['statement_section'] or ('operating_revenue' if r['type'] == 'income' else 'operating_expense')
        sections[sec] = sections.get(sec, 0) + float(r['amount'])

    operating_revenue = sections.get('operating_revenue', 0)
    other_revenue     = sections.get('other_revenue', 0)
    cogs              = sections.get('cogs', 0)
    operating_expense = sections.get('operating_expense', 0)
    other_expense     = sections.get('other_expense', 0)
    total_revenue     = operating_revenue + other_revenue
    gross_profit      = operating_revenue - cogs
    operating_income  = gross_profit - operating_expense
    net_income        = operating_income + other_revenue - other_expense

    income_stmt = {
        'operating_revenue': round(operating_revenue, 2),
        'other_revenue':     round(other_revenue, 2),
        'total_revenue':     round(total_revenue, 2),
        'cogs':              round(cogs, 2),
        'gross_profit':      round(gross_profit, 2),
        'operating_expense': round(operating_expense, 2),
        'operating_income':  round(operating_income, 2),
        'other_expense':     round(other_expense, 2),
        'net_income':        round(net_income, 2),
    }

    total_expense = cogs + operating_expense + other_expense
    cash = opening_cash + total_revenue - total_expense
    equity = opening_equity + net_income

    balance_sheet = {
        'cash':            round(cash, 2),
        'total_assets':    round(cash, 2),
        'total_liabilities': 0,
        'equity':          round(equity, 2),
    }

    cash_flow = {
        'operating': round(net_income, 2),
        'investing':  0,
        'financing':  0,
        'net_change': round(net_income, 2),
        'ending_cash': round(cash, 2),
    }

    return income_stmt, balance_sheet, cash_flow


# ── Finance Categories ─────────────────────────────────────────

@bp.route('/api/finance/categories', methods=['GET'])
@require_module('finance')
def api_finance_categories_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM finance_categories ORDER BY sort_order, id").fetchall()
    return jsonify([_finance_cat_row(r) for r in rows])


@bp.route('/api/finance/categories', methods=['POST'])
@require_module('finance')
def api_finance_category_create():
    b = request.get_json(force=True)
    if not b.get('name', '').strip(): return jsonify({'error': '名稱為必填'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_categories (name,type,color,sort_order,active,statement_section)
            VALUES (%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b['name'].strip(), b.get('type', 'expense'), b.get('color', '#4a7bda'),
              int(b.get('sort_order', 0)), bool(b.get('active', True)),
              b.get('statement_section') or ('operating_revenue' if b.get('type') == 'income' else 'operating_expense')
              )).fetchone()
    return jsonify(_finance_cat_row(row)), 201


@bp.route('/api/finance/categories/<int:cid>', methods=['PUT'])
@require_module('finance')
def api_finance_category_update(cid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_categories SET name=%s,type=%s,color=%s,sort_order=%s,active=%s,statement_section=%s
            WHERE id=%s RETURNING *
        """, (b.get('name', '').strip(), b.get('type', 'expense'), b.get('color', '#4a7bda'),
              int(b.get('sort_order', 0)), bool(b.get('active', True)),
              b.get('statement_section') or ('operating_revenue' if b.get('type') == 'income' else 'operating_expense'),
              cid)).fetchone()
    return jsonify(_finance_cat_row(row)) if row else ('', 404)


@bp.route('/api/finance/categories/<int:cid>', methods=['DELETE'])
@require_module('finance')
def api_finance_category_delete(cid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_categories WHERE id=%s", (cid,))
    return jsonify({'deleted': cid})


# ── Finance Records ────────────────────────────────────────────

@bp.route('/api/finance/records', methods=['GET'])
@require_module('finance')
def api_finance_records_list():
    month  = request.args.get('month', '')
    ftype  = request.args.get('type', '')
    cat_id = request.args.get('category_id', '')
    conds, params = ['TRUE'], []
    if month:   conds.append("to_char(fr.record_date,'YYYY-MM')=%s"); params.append(month)
    if ftype:   conds.append("fr.type=%s"); params.append(ftype)
    if cat_id:  conds.append("fr.category_id=%s"); params.append(int(cat_id))
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT fr.*, fc.name as category_name, fc.color as category_color,
                   fd.filename as doc_filename, fd.ocr_raw as ocr_raw
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            LEFT JOIN finance_documents fd ON fd.id=fr.document_id
            WHERE {' AND '.join(conds)}
            ORDER BY fr.record_date DESC, fr.id DESC
        """, params).fetchall()
    result = []
    for r in rows:
        d = _finance_rec_row(r)
        d['category_name']  = r['category_name']
        d['category_color'] = r['category_color']
        d['doc_filename']   = r['doc_filename']
        d['ocr_raw']        = r['ocr_raw'] if r['ocr_raw'] else None
        result.append(d)
    return jsonify(result)


@bp.route('/api/finance/documents', methods=['GET'])
@require_module('finance')
def api_finance_documents_list():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT fd.*,
                   COUNT(fr.id) as linked_count,
                   MAX(fr.title) as linked_title,
                   MAX(fr.id) as linked_record_id
            FROM finance_documents fd
            LEFT JOIN finance_records fr ON fr.document_id = fd.id
            GROUP BY fd.id
            ORDER BY fd.created_at DESC
        """).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if d.get('upload_date'): d['upload_date'] = str(d['upload_date'])
        if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
        d['linked_count'] = int(d['linked_count'] or 0)
        result.append(d)
    return jsonify(result)


@bp.route('/api/finance/records', methods=['POST'])
@require_module('finance')
def api_finance_record_create():
    b = request.get_json(force=True)
    if not b.get('title', '').strip(): return jsonify({'error': '標題為必填'}), 400
    if not b.get('record_date'):       return jsonify({'error': '日期為必填'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_records
              (record_date, category_id, type, title, amount, tax_amount,
               vendor, invoice_no, note, document_id, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b['record_date'], b.get('category_id') or None, b.get('type', 'expense'),
              b['title'].strip(), float(b.get('amount', 0)), float(b.get('tax_amount', 0)),
              b.get('vendor', '').strip(), b.get('invoice_no', '').strip(),
              b.get('note', '').strip(), b.get('document_id') or None,
              session.get('admin_display_name', ''))).fetchone()
    return jsonify(_finance_rec_row(row)), 201


@bp.route('/api/finance/records/<int:rid>', methods=['PUT'])
@require_module('finance')
def api_finance_record_update(rid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_records SET
              record_date=%s, category_id=%s, type=%s, title=%s, amount=%s,
              tax_amount=%s, vendor=%s, invoice_no=%s, note=%s, updated_at=NOW()
            WHERE id=%s RETURNING *
        """, (b['record_date'], b.get('category_id') or None, b.get('type', 'expense'),
              b.get('title', '').strip(), float(b.get('amount', 0)), float(b.get('tax_amount', 0)),
              b.get('vendor', '').strip(), b.get('invoice_no', '').strip(),
              b.get('note', '').strip(), rid)).fetchone()
    return jsonify(_finance_rec_row(row)) if row else ('', 404)


@bp.route('/api/finance/records/<int:rid>', methods=['DELETE'])
@require_module('finance')
def api_finance_record_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_records WHERE id=%s", (rid,))
    return jsonify({'deleted': rid})


# ── Finance P&L Summary ────────────────────────────────────────

@bp.route('/api/finance/summary/<year>/<month>', methods=['GET'])
@require_module('finance')
def api_finance_summary(year, month):
    period = f"{year}-{month.zfill(2)}"
    with get_db() as conn:
        totals = conn.execute("""
            SELECT type, COALESCE(SUM(amount),0) as total
            FROM finance_records
            WHERE to_char(record_date,'YYYY-MM')=%s
            GROUP BY type
        """, (period,)).fetchall()
        income  = next((float(r['total']) for r in totals if r['type'] == 'income'), 0.0)
        expense = next((float(r['total']) for r in totals if r['type'] == 'expense'), 0.0)

        by_cat = conn.execute("""
            SELECT fc.name, fc.color, fr.type, COALESCE(SUM(fr.amount),0) as total
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            WHERE to_char(fr.record_date,'YYYY-MM')=%s
            GROUP BY fc.name, fc.color, fr.type
            ORDER BY total DESC
        """, (period,)).fetchall()

        trend = conn.execute("""
            SELECT to_char(record_date,'YYYY-MM') as mon,
                   type, COALESCE(SUM(amount),0) as total
            FROM finance_records
            WHERE record_date >= (DATE_TRUNC('month', %s::date) - INTERVAL '5 months')
              AND record_date <  (DATE_TRUNC('month', %s::date) + INTERVAL '1 month')
            GROUP BY to_char(record_date,'YYYY-MM'), type
            ORDER BY mon
        """, (f"{period}-01", f"{period}-01")).fetchall()

    return jsonify({
        'income':  income,
        'expense': expense,
        'net':     income - expense,
        'by_category': [
            {'name': r['name'] or '未分類', 'color': r['color'] or '#8892a4',
             'type': r['type'], 'total': float(r['total'])}
            for r in by_cat
        ],
        'trend': [
            {'month': r['mon'], 'type': r['type'], 'total': float(r['total'])}
            for r in trend
        ],
    })


# ── Finance OCR ────────────────────────────────────────────────

@bp.route('/api/finance/ocr', methods=['POST'])
@require_module('finance')
def api_finance_ocr():
    import anthropic as _ant, base64, re as _re
    from config import ANTHROPIC_API_KEY
    if not ANTHROPIC_API_KEY:
        return jsonify({'error': '尚未設定 ANTHROPIC_API_KEY 環境變數'}), 500
    file = request.files.get('file')
    if not file: return jsonify({'error': '請上傳圖片或 PDF 檔案'}), 400
    raw = file.read()
    media_type = file.content_type or 'image/jpeg'
    if media_type not in ('image/jpeg', 'image/png', 'image/gif', 'image/webp'):
        media_type = 'image/jpeg'
    img_b64 = base64.standard_b64encode(raw).decode()
    client = _ant.Anthropic(api_key=ANTHROPIC_API_KEY)
    try:
        msg = client.messages.create(
            model='claude-sonnet-4-6', max_tokens=1024,
            messages=[{'role': 'user', 'content': [
                {'type': 'image', 'source': {'type': 'base64', 'media_type': media_type, 'data': img_b64}},
                {'type': 'text', 'text': (
                    '請辨識此文件，以JSON格式回傳以下欄位（找不到的欄位填null）：\n'
                    '{"date":"YYYY-MM-DD","vendor":"廠商名稱","invoice_no":"發票或單據號碼",'
                    '"total_amount":含稅總金額數字,"tax_amount":稅額數字,"pre_tax_amount":未稅金額數字,'
                    '"doc_type":"invoice或receipt或expense之一",'
                    '"title":"建議記帳標題（簡短）",'
                    '"items":[{"name":"品項","qty":數量,"unit_price":單價,"amount":小計}],'
                    '"currency":"TWD"}\n只回傳JSON，不要其他文字或markdown。'
                )}
            ]}]
        )
        text = msg.content[0].text.strip()
        text = _re.sub(r'^```json\s*', '', text, flags=_re.MULTILINE)
        text = _re.sub(r'\s*```$', '', text, flags=_re.MULTILINE)
        result = _json.loads(text)
    except _json.JSONDecodeError:
        result = {'raw_text': text, 'error': 'OCR 回傳格式無法解析'}
    except Exception as e:
        return jsonify({'error': f'OCR 失敗：{str(e)}'}), 500
    try:
        with get_db() as conn:
            doc = conn.execute("""
                INSERT INTO finance_documents (filename, doc_type, ocr_raw, upload_date)
                VALUES (%s,%s,%s,CURRENT_DATE) RETURNING id
            """, (file.filename, result.get('doc_type', ''), _json.dumps(result))).fetchone()
        result['document_id'] = doc['id']
    except Exception as e:
        print(f"[finance_ocr doc save] {e}")
    return jsonify(result)


# ── Finance Export ─────────────────────────────────────────────

@bp.route('/api/finance/export', methods=['GET'])
@require_module('finance')
def api_finance_export():
    month = request.args.get('month', '')
    conds, params = ['TRUE'], []
    if month:
        conds.append("to_char(fr.record_date,'YYYY-MM')=%s"); params.append(month)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT fr.record_date, fr.type, fr.title, fr.amount, fr.tax_amount,
                   fr.vendor, fr.invoice_no, fr.note, fc.name as category_name
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            WHERE {' AND '.join(conds)}
            ORDER BY fr.record_date, fr.id
        """, params).fetchall()
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'財務記錄_{month or "全部"}'
    thin = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
                  top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
    hfill = PatternFill('solid', fgColor='0F1C3A')
    ca = Alignment(horizontal='center', vertical='center')
    headers = ['日期', '類型', '類別', '標題', '金額', '稅額', '廠商', '單據號碼', '備註']
    col_w = [12, 6, 12, 20, 12, 10, 14, 14, 20]
    for ci, (h, w2) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill = hfill; cell.alignment = ca; cell.border = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w2
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    income_fill  = PatternFill('solid', fgColor='E8F5E9')
    expense_fill = PatternFill('solid', fgColor='FDECEA')
    for ri, r in enumerate(rows, 2):
        is_income = r['type'] == 'income'
        vals = [str(r['record_date']), '收入' if is_income else '支出',
                r['category_name'] or '', r['title'],
                float(r['amount']), float(r['tax_amount'] or 0),
                r['vendor'] or '', r['invoice_no'] or '', r['note'] or '']
        row_fill = income_fill if is_income else expense_fill
        for ci2, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci2, value=v)
            cell.fill = row_fill
            cell.alignment = ca if ci2 != 9 else Alignment(vertical='center')
            cell.border = thin
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from flask import Response
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=finance_{month or "all"}.xlsx'})


# ── Finance Settings ───────────────────────────────────────────

@bp.route('/api/finance/settings', methods=['GET'])
@require_module('finance')
def api_finance_settings_get():
    return jsonify(_get_finance_settings())


@bp.route('/api/finance/settings', methods=['POST'])
@require_module('finance')
def api_finance_settings_save():
    b = request.get_json(force=True) or {}
    with get_db() as conn:
        for k, v in b.items():
            conn.execute("""
                INSERT INTO finance_settings (setting_key, setting_value)
                VALUES (%s,%s)
                ON CONFLICT (setting_key) DO UPDATE SET setting_value=EXCLUDED.setting_value
            """, (k, str(v)))
    return jsonify({'ok': True})


# ── Financial Statements ───────────────────────────────────────

@bp.route('/api/finance/statements', methods=['GET'])
@require_module('finance')
def api_finance_statements():
    year  = int(request.args.get('year',  0) or 0)
    month = int(request.args.get('month', 0) or 0)
    if not year or not month:
        from datetime import datetime, timezone, timedelta
        now = datetime.now(timezone(timedelta(hours=8)))
        year, month = now.year, now.month
    income_stmt, balance_sheet, cash_flow = _compute_statements(year, month)
    roc_year = year - 1911
    return jsonify({
        'year': year, 'month': month, 'roc_year': roc_year,
        'income_statement': income_stmt,
        'balance_sheet':    balance_sheet,
        'cash_flow':        cash_flow,
    })


@bp.route('/api/finance/statements/export', methods=['GET'])
@require_module('finance')
def api_finance_statements_export():
    year  = int(request.args.get('year',  0) or 0)
    month = int(request.args.get('month', 0) or 0)
    if not year or not month:
        from datetime import datetime, timezone, timedelta
        now = datetime.now(timezone(timedelta(hours=8)))
        year, month = now.year, now.month
    income_stmt, balance_sheet, cash_flow = _compute_statements(year, month)
    settings = _get_finance_settings()
    company  = settings.get('company_name', '')
    roc_year = year - 1911
    period_str = f"民國{roc_year}年{month}月"

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    wb = openpyxl.Workbook()

    hfill = PatternFill('solid', fgColor='0F1C3A')
    thin  = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
                   top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
    ca    = Alignment(horizontal='center', vertical='center')
    ra    = Alignment(horizontal='right', vertical='center')
    hfont = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)

    def _sheet(ws, title, items):
        ws.title = title
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 16
        ws.merge_cells('A1:B1')
        c = ws['A1']
        c.value = f'{company} {period_str} {title}'
        c.font = hfont; c.fill = hfill; c.alignment = ca
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A3'
        for ci2, h in enumerate(['項目', '金額（NT$）'], 1):
            cell = ws.cell(row=2, column=ci2, value=h)
            cell.font = hfont; cell.fill = hfill; cell.alignment = ca; cell.border = thin
        row = 3
        for label, val in items:
            ws.cell(row=row, column=1, value=label).border = thin
            c2 = ws.cell(row=row, column=2, value=val)
            c2.alignment = ra; c2.border = thin
            if isinstance(val, (int, float)):
                c2.number_format = '#,##0.00'
            row += 1

    _sheet(wb.active, '損益表', [
        ('營業收入', income_stmt['operating_revenue']),
        ('其他收入', income_stmt['other_revenue']),
        ('收入合計', income_stmt['total_revenue']),
        ('', ''),
        ('銷貨成本', income_stmt['cogs']),
        ('毛利', income_stmt['gross_profit']),
        ('營業費用', income_stmt['operating_expense']),
        ('營業利益', income_stmt['operating_income']),
        ('其他費用', income_stmt['other_expense']),
        ('本期淨利', income_stmt['net_income']),
    ])
    ws2 = wb.create_sheet()
    _sheet(ws2, '資產負債表', [
        ('現金及約當現金', balance_sheet['cash']),
        ('資產合計', balance_sheet['total_assets']),
        ('', ''),
        ('負債合計', balance_sheet['total_liabilities']),
        ('業主權益', balance_sheet['equity']),
        ('負債+權益合計', balance_sheet['total_liabilities'] + balance_sheet['equity']),
    ])
    ws3 = wb.create_sheet()
    _sheet(ws3, '現金流量表', [
        ('營業活動現金流量', cash_flow['operating']),
        ('投資活動現金流量', cash_flow['investing']),
        ('籌資活動現金流量', cash_flow['financing']),
        ('現金淨增減', cash_flow['net_change']),
        ('期末現金', cash_flow['ending_cash']),
    ])

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from flask import Response
    fn = f"statements_{roc_year}_{month:02d}.xlsx"
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={fn}'})


# ── Recurring Entries ──────────────────────────────────────────

def _recurring_row(r):
    if not r: return None
    d = dict(r)
    if d.get('start_date'): d['start_date'] = str(d['start_date'])
    if d.get('end_date'):   d['end_date']   = str(d['end_date'])
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    if d.get('amount') is not None:     d['amount']     = float(d['amount'])
    if d.get('tax_amount') is not None: d['tax_amount'] = float(d['tax_amount'])
    return d


@bp.route('/api/finance/recurring', methods=['GET'])
@require_module('finance')
def api_recurring_list():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT fr.*, fc.name as category_name
            FROM finance_recurring fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            ORDER BY fr.id DESC
        """).fetchall()
    result = []
    for r in rows:
        d = _recurring_row(r)
        d['category_name'] = r['category_name']
        result.append(d)
    return jsonify(result)


@bp.route('/api/finance/recurring', methods=['POST'])
@require_module('finance')
def api_recurring_create():
    b = request.get_json(force=True)
    if not b.get('title', '').strip(): return jsonify({'error': '標題為必填'}), 400
    if not b.get('start_date'):        return jsonify({'error': '開始日期為必填'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_recurring
              (title, type, category_id, amount, tax_amount, vendor, note,
               frequency, day_of_month, start_date, end_date, active)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b['title'].strip(), b.get('type', 'expense'), b.get('category_id') or None,
              float(b.get('amount', 0)), float(b.get('tax_amount', 0)),
              b.get('vendor', '').strip(), b.get('note', '').strip(),
              b.get('frequency', 'monthly'), int(b.get('day_of_month', 1)),
              b['start_date'], b.get('end_date') or None,
              bool(b.get('active', True)))).fetchone()
    return jsonify(_recurring_row(row)), 201


@bp.route('/api/finance/recurring/<int:rid>', methods=['PUT'])
@require_module('finance')
def api_recurring_update(rid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_recurring SET
              title=%s, type=%s, category_id=%s, amount=%s, tax_amount=%s,
              vendor=%s, note=%s, frequency=%s, day_of_month=%s,
              start_date=%s, end_date=%s, active=%s
            WHERE id=%s RETURNING *
        """, (b.get('title', '').strip(), b.get('type', 'expense'), b.get('category_id') or None,
              float(b.get('amount', 0)), float(b.get('tax_amount', 0)),
              b.get('vendor', '').strip(), b.get('note', '').strip(),
              b.get('frequency', 'monthly'), int(b.get('day_of_month', 1)),
              b.get('start_date'), b.get('end_date') or None,
              bool(b.get('active', True)), rid)).fetchone()
    return jsonify(_recurring_row(row)) if row else ('', 404)


@bp.route('/api/finance/recurring/<int:rid>', methods=['DELETE'])
@require_module('finance')
def api_recurring_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_recurring WHERE id=%s", (rid,))
    return jsonify({'deleted': rid})


@bp.route('/api/finance/recurring/generate', methods=['POST'])
@require_module('finance')
def api_recurring_generate():
    from datetime import date as _d
    b     = request.get_json(force=True) or {}
    month = (b.get('month') or '').strip()
    if not month:
        month = _d.today().strftime('%Y-%m')
    try:
        y, mo = int(month[:4]), int(month[5:7])
    except Exception:
        return jsonify({'error': '月份格式錯誤'}), 400

    import calendar as _cal
    days_in = _cal.monthrange(y, mo)[1]
    created = 0

    with get_db() as conn:
        templates = conn.execute(
            "SELECT * FROM finance_recurring WHERE active=TRUE"
        ).fetchall()
        for t in templates:
            start = _d.fromisoformat(str(t['start_date']))
            end   = _d.fromisoformat(str(t['end_date'])) if t['end_date'] else None
            target_month_start = _d(y, mo, 1)
            target_month_end   = _d(y, mo, days_in)
            if start > target_month_end: continue
            if end and end < target_month_start: continue

            freq = t['frequency']
            if freq == 'monthly':
                dom = min(int(t['day_of_month']), days_in)
                record_date = _d(y, mo, dom)
            elif freq == 'quarterly':
                if (mo - (start.month - 1)) % 3 != 0: continue
                dom = min(int(t['day_of_month']), days_in)
                record_date = _d(y, mo, dom)
            elif freq == 'yearly':
                if mo != start.month: continue
                dom = min(int(t['day_of_month']), days_in)
                record_date = _d(y, mo, dom)
            else:
                continue

            existing = conn.execute("""
                SELECT id FROM finance_records
                WHERE title=%s AND record_date=%s AND created_by='recurring'
            """, (t['title'], str(record_date))).fetchone()
            if existing: continue

            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, tax_amount,
                   vendor, note, created_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'recurring')
            """, (str(record_date), t['category_id'], t['type'],
                  t['title'], t['amount'], t['tax_amount'] or 0,
                  t['vendor'] or '', t['note'] or ''))
            conn.execute(
                "UPDATE finance_recurring SET last_generated=%s WHERE id=%s",
                (month, t['id'])
            )
            created += 1
    return jsonify({'ok': True, 'month': month, 'created': created})


# ── Bank Statements ────────────────────────────────────────────

@bp.route('/api/finance/bank/import', methods=['POST'])
@require_module('finance')
def api_bank_import():
    import csv, io as _io
    from datetime import date as _d
    file = request.files.get('file')
    if not file: return jsonify({'error': '請上傳 CSV 檔案'}), 400
    content = file.read().decode('utf-8-sig', errors='replace')
    reader  = csv.DictReader(_io.StringIO(content))
    rows    = list(reader)
    if not rows: return jsonify({'error': 'CSV 無資料'}), 400

    headers = [h.strip() for h in (reader.fieldnames or [])]
    date_col   = next((h for h in headers if '日期' in h or 'date' in h.lower()), headers[0] if headers else '')
    amount_col = next((h for h in headers if '金額' in h or 'amount' in h.lower()), None)
    debit_col  = next((h for h in headers if '支出' in h or '借' in h or 'debit' in h.lower()), None)
    credit_col = next((h for h in headers if '存入' in h or '貸' in h or 'credit' in h.lower()), None)
    desc_col   = next((h for h in headers if '摘要' in h or '說明' in h or 'desc' in h.lower()), None)
    acct_col   = next((h for h in headers if '帳號' in h or 'account' in h.lower()), None)

    account_name = request.form.get('account_name', '').strip()
    import_batch = _d.today().isoformat()
    imported = 0

    with get_db() as conn:
        for row in rows:
            try:
                date_str = row.get(date_col, '').strip()
                if not date_str: continue
                if '/' in date_str:
                    parts = date_str.split('/')
                    if len(parts) == 3 and len(parts[0]) <= 4 and int(parts[0]) < 200:
                        y = int(parts[0]) + 1911
                        date_str = f"{y}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
                    else:
                        parts2 = [p.zfill(2) for p in parts]
                        date_str = '-'.join(parts2) if len(parts) == 3 else date_str

                if amount_col:
                    raw_amt = row.get(amount_col, '0').replace(',', '').strip()
                    amount  = abs(float(raw_amt)) if raw_amt else 0
                    txn_type = 'credit' if amount >= 0 else 'debit'
                else:
                    debit  = abs(float((row.get(debit_col,  '0') or '0').replace(',', '')))
                    credit = abs(float((row.get(credit_col, '0') or '0').replace(',', '')))
                    if credit > 0:
                        amount = credit; txn_type = 'credit'
                    else:
                        amount = debit; txn_type = 'debit'

                if amount == 0: continue
                description = row.get(desc_col, '').strip() if desc_col else ''
                acct_name   = row.get(acct_col, account_name).strip() if acct_col else account_name

                conn.execute("""
                    INSERT INTO bank_statements
                      (account_name, txn_date, amount, txn_type, description, import_batch)
                    VALUES (%s,%s,%s,%s,%s,%s)
                """, (acct_name, date_str, amount, txn_type, description, import_batch))
                imported += 1
            except Exception as e:
                print(f"[bank_import row] {e}")

    return jsonify({'ok': True, 'imported': imported, 'batch': import_batch})


@bp.route('/api/finance/bank/statements', methods=['GET'])
@require_module('finance')
def api_bank_statements():
    month = request.args.get('month', '')
    conds, params = ['TRUE'], []
    if month:
        conds.append("TO_CHAR(bs.txn_date,'YYYY-MM')=%s"); params.append(month)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT bs.*, fr.title as matched_title, fr.amount as matched_amount,
                   fr.record_date as matched_date
            FROM bank_statements bs
            LEFT JOIN finance_records fr ON fr.id=bs.matched_record_id
            WHERE {' AND '.join(conds)}
            ORDER BY bs.txn_date DESC, bs.id DESC
        """, params).fetchall()
    result = []
    for r in rows:
        d = _bank_row(r)
        d['matched_title']  = r['matched_title']
        d['matched_amount'] = float(r['matched_amount']) if r['matched_amount'] else None
        d['matched_date']   = str(r['matched_date']) if r['matched_date'] else None
        result.append(d)
    return jsonify(result)


@bp.route('/api/finance/bank/statements/<int:sid>', methods=['DELETE'])
@require_module('finance')
def api_bank_statement_delete(sid):
    with get_db() as conn:
        conn.execute("DELETE FROM bank_statements WHERE id=%s", (sid,))
    return jsonify({'deleted': sid})


@bp.route('/api/finance/bank/match', methods=['POST'])
@require_module('finance')
def api_bank_match():
    b   = request.get_json(force=True)
    sid = b.get('statement_id')
    rid = b.get('record_id')
    with get_db() as conn:
        if rid:
            conn.execute(
                "UPDATE bank_statements SET reconciled=TRUE, matched_record_id=%s WHERE id=%s",
                (rid, sid))
        else:
            conn.execute(
                "UPDATE bank_statements SET reconciled=FALSE, matched_record_id=NULL WHERE id=%s",
                (sid,))
    return jsonify({'ok': True})


@bp.route('/api/finance/bank/auto-match', methods=['POST'])
@require_module('finance')
def api_bank_auto_match():
    b     = request.get_json(force=True)
    month = b.get('month', '')
    matched = 0
    with get_db() as conn:
        stmts = conn.execute(
            "SELECT * FROM bank_statements WHERE reconciled=FALSE"
            + (" AND TO_CHAR(txn_date,'YYYY-MM')=%s" if month else ""),
            ([month] if month else [])
        ).fetchall()
        for s in stmts:
            ftype = 'income' if s['txn_type'] == 'credit' else 'expense'
            rec = conn.execute("""
                SELECT id FROM finance_records
                WHERE type=%s AND amount=%s
                  AND ABS(record_date - %s::date) <= 3
                  AND id NOT IN (
                      SELECT matched_record_id FROM bank_statements
                      WHERE matched_record_id IS NOT NULL
                  )
                ORDER BY ABS(record_date - %s::date), id
                LIMIT 1
            """, (ftype, s['amount'], s['txn_date'], s['txn_date'])).fetchone()
            if rec:
                conn.execute(
                    "UPDATE bank_statements SET reconciled=TRUE, matched_record_id=%s WHERE id=%s",
                    (rec['id'], s['id'])
                )
                matched += 1
    return jsonify({'matched': matched})


@bp.route('/api/finance/bank/summary', methods=['GET'])
@require_module('finance')
def api_bank_summary():
    month  = request.args.get('month', '')
    cond   = "AND TO_CHAR(txn_date,'YYYY-MM')=%s" if month else ""
    params = [month] if month else []
    with get_db() as conn:
        r = conn.execute(f"""
            SELECT
              COUNT(*) as total,
              SUM(CASE WHEN reconciled THEN 1 ELSE 0 END) as matched,
              SUM(CASE WHEN txn_type='credit' THEN amount ELSE 0 END) as total_credit,
              SUM(CASE WHEN txn_type='debit'  THEN amount ELSE 0 END) as total_debit,
              SUM(CASE WHEN reconciled AND txn_type='credit' THEN amount ELSE 0 END) as matched_credit,
              SUM(CASE WHEN reconciled AND txn_type='debit'  THEN amount ELSE 0 END) as matched_debit
            FROM bank_statements WHERE TRUE {cond}
        """, params).fetchone()
    d = dict(r)
    for k in d:
        if d[k] is not None:
            try: d[k] = float(d[k])
            except: d[k] = int(d[k]) if d[k] is not None else None
    return jsonify(d)


# ── Tax VAT 401 ────────────────────────────────────────────────

@bp.route('/api/finance/tax/<int:year>/<int:period>', methods=['GET'])
@require_module('finance')
def api_finance_tax(year, period):
    if period < 1 or period > 6:
        return jsonify({'error': '期別需為 1-6'}), 400
    m_start = (period - 1) * 2 + 1
    m_end   = m_start + 1
    months  = [f"{year}-{str(m).zfill(2)}" for m in range(m_start, m_end + 1)]

    with get_db() as conn:
        rows = conn.execute("""
            SELECT fr.type, fr.amount, fr.tax_amount, fr.title,
                   fr.vendor, fr.invoice_no, fr.record_date,
                   fc.name as category_name
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            WHERE TO_CHAR(fr.record_date,'YYYY-MM') = ANY(%s)
            ORDER BY fr.record_date, fr.type
        """, (months,)).fetchall()

    sales_rows    = [r for r in rows if r['type'] == 'income']
    purchase_rows = [r for r in rows if r['type'] == 'expense']
    sales_amount    = sum(float(r['amount'])           for r in sales_rows)
    sales_tax       = sum(float(r['tax_amount'] or 0)  for r in sales_rows)
    purchase_amount = sum(float(r['amount'])           for r in purchase_rows)
    purchase_tax    = sum(float(r['tax_amount'] or 0)  for r in purchase_rows)
    tax_payable     = round(sales_tax - purchase_tax, 2)

    def _fmt(r):
        return {
            'date': str(r['record_date']), 'title': r['title'],
            'vendor': r['vendor'] or '', 'invoice_no': r['invoice_no'] or '',
            'amount': float(r['amount']), 'tax_amount': float(r['tax_amount'] or 0),
            'category': r['category_name'] or '未分類',
        }
    return jsonify({
        'year': year, 'period': period, 'roc_year': year - 1911,
        'months': months,
        'sales':     {'rows': [_fmt(r) for r in sales_rows],     'amount': round(sales_amount, 2),    'tax': round(sales_tax, 2)},
        'purchases': {'rows': [_fmt(r) for r in purchase_rows],  'amount': round(purchase_amount, 2), 'tax': round(purchase_tax, 2)},
        'tax_payable': tax_payable,
        'is_refund':   tax_payable < 0,
    })


@bp.route('/api/finance/tax/<int:year>/<int:period>/sync', methods=['POST'])
@require_module('finance')
def api_finance_tax_sync(year, period):
    if period < 1 or period > 6:
        return jsonify({'error': '期別需為 1-6'}), 400
    m_start  = (period - 1) * 2 + 1
    m_end    = m_start + 1
    months   = [f"{year}-{str(m).zfill(2)}" for m in range(m_start, m_end + 1)]
    roc_year = year - 1911

    with get_db() as conn:
        rows = conn.execute("""
            SELECT type, SUM(tax_amount) as tax_total
            FROM finance_records
            WHERE TO_CHAR(record_date,'YYYY-MM') = ANY(%s)
              AND tax_amount IS NOT NULL AND tax_amount <> 0
            GROUP BY type
        """, (months,)).fetchall()

    sales_tax    = sum(float(r['tax_total']) for r in rows if r['type'] == 'income')
    purchase_tax = sum(float(r['tax_total']) for r in rows if r['type'] == 'expense')
    tax_payable  = round(sales_tax - purchase_tax, 2)
    if tax_payable == 0:
        return jsonify({'created': 0, 'message': '稅額為零，無需建立分錄'})

    import calendar as _cal
    record_date  = f"{year}-{str(m_end).zfill(2)}-{_cal.monthrange(year, m_end)[1]}"
    note         = f"銷項稅 ${round(sales_tax,0):,.0f} − 進項稅 ${round(purchase_tax,0):,.0f} = {'應繳' if tax_payable>0 else '退稅'} ${abs(round(tax_payable,0)):,.0f}"
    period_label = f"民國{roc_year}年第{period}期（{months[0]}～{months[-1]}）"

    created = 0
    with get_db() as conn:
        if tax_payable > 0:
            cat = conn.execute(
                "SELECT id FROM finance_categories WHERE name='稅費' AND type='expense' LIMIT 1"
            ).fetchone()
            if not cat:
                cat = conn.execute("""
                    INSERT INTO finance_categories (name, type, color, sort_order, statement_section)
                    VALUES ('稅費','expense','#8892a4', 99,'operating_expense') RETURNING *
                """).fetchone()
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, tax_amount, note, created_by)
                VALUES (%s,%s,'expense',%s,%s,0,%s,'tax-sync')
            """, (record_date, cat['id'], f"應繳營業稅 {period_label}", tax_payable, note))
            created += 1
        else:
            cat = conn.execute(
                "SELECT id FROM finance_categories WHERE name='其他收入' AND type='income' LIMIT 1"
            ).fetchone()
            if not cat:
                cat = conn.execute("""
                    INSERT INTO finance_categories (name, type, color, sort_order, statement_section)
                    VALUES ('其他收入','income','#c8a96e', 99,'other_revenue') RETURNING *
                """).fetchone()
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, tax_amount, note, created_by)
                VALUES (%s,%s,'income',%s,%s,0,%s,'tax-sync')
            """, (record_date, cat['id'], f"營業稅退稅 {period_label}", abs(tax_payable), note))
            created += 1
    return jsonify({'created': created, 'tax_payable': tax_payable, 'record_date': record_date})


# ── AR/AP Tracking ─────────────────────────────────────────────

@bp.route('/api/finance/payables', methods=['GET'])
@require_module('finance')
def api_payables_list():
    from datetime import date as _d
    ptype  = request.args.get('type', '')
    status = request.args.get('status', '')
    conds, params = ['TRUE'], []
    if ptype:  conds.append("payable_type=%s"); params.append(ptype)
    if status == 'overdue':
        conds.append("status='open' AND due_date < CURRENT_DATE")
    elif status:
        conds.append("status=%s"); params.append(status)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT *, CURRENT_DATE - due_date AS days_overdue
            FROM finance_payables
            WHERE {' AND '.join(conds)}
            ORDER BY
              CASE WHEN status='open' AND due_date < CURRENT_DATE THEN 0
                   WHEN status='open' THEN 1
                   ELSE 2 END,
              due_date
        """, params).fetchall()
    result = []
    for r in rows:
        d = _payable_row(r)
        d['days_overdue'] = int(r['days_overdue']) if r['days_overdue'] is not None else 0
        if d['status'] == 'open' and d.get('due_date') and str(_d.today()) > d['due_date']:
            d['effective_status'] = 'overdue'
        else:
            d['effective_status'] = d['status']
        result.append(d)
    return jsonify(result)


@bp.route('/api/finance/payables', methods=['POST'])
@require_module('finance')
def api_payable_create():
    b = request.get_json(force=True)
    if not b.get('title', '').strip(): return jsonify({'error': '標題為必填'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_payables
              (payable_type, title, party_name, invoice_no, amount, due_date, status, note)
            VALUES (%s,%s,%s,%s,%s,%s,'open',%s) RETURNING *
        """, (b.get('payable_type', 'payable'), b['title'].strip(),
              b.get('party_name', '').strip(), b.get('invoice_no', '').strip(),
              float(b.get('amount', 0)), b.get('due_date') or None,
              b.get('note', '').strip())).fetchone()
    return jsonify(_payable_row(row)), 201


@bp.route('/api/finance/payables/<int:pid>', methods=['PUT'])
@require_module('finance')
def api_payable_update(pid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_payables SET
              payable_type=%s, title=%s, party_name=%s, invoice_no=%s,
              amount=%s, due_date=%s, status=%s, paid_date=%s,
              note=%s, updated_at=NOW()
            WHERE id=%s RETURNING *
        """, (b.get('payable_type', 'payable'), b.get('title', '').strip(),
              b.get('party_name', '').strip(), b.get('invoice_no', '').strip(),
              float(b.get('amount', 0)), b.get('due_date') or None,
              b.get('status', 'open'), b.get('paid_date') or None,
              b.get('note', '').strip(), pid)).fetchone()
    return jsonify(_payable_row(row)) if row else ('', 404)


@bp.route('/api/finance/payables/<int:pid>', methods=['DELETE'])
@require_module('finance')
def api_payable_delete(pid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_payables WHERE id=%s", (pid,))
    return jsonify({'deleted': pid})


@bp.route('/api/finance/payables/aging', methods=['GET'])
@require_module('finance')
def api_payables_aging():
    ptype = request.args.get('type', 'payable')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT *, CURRENT_DATE - due_date AS days_overdue
            FROM finance_payables
            WHERE payable_type=%s AND status='open'
        """, (ptype,)).fetchall()
    buckets      = {'current': 0, 'd1_30': 0, 'd31_60': 0, 'd61_90': 0, 'd90plus': 0}
    bucket_rows  = {'current': [], 'd1_30': [], 'd31_60': [], 'd61_90': [], 'd90plus': []}
    for r in rows:
        do = int(r['days_overdue']) if r['days_overdue'] is not None else 0
        d  = _payable_row(r); d['days_overdue'] = do
        k  = 'current' if do <= 0 else ('d1_30' if do <= 30 else ('d31_60' if do <= 60 else ('d61_90' if do <= 90 else 'd90plus')))
        buckets[k]     += float(r['amount'])
        bucket_rows[k].append(d)
    return jsonify({'buckets': buckets, 'rows': bucket_rows, 'type': ptype})


# ── Budget Management ──────────────────────────────────────────

@bp.route('/api/finance/budgets', methods=['GET'])
@require_module('finance')
def api_budgets_list():
    year  = request.args.get('year',  '')
    month = request.args.get('month', '')
    if not year or not month:
        from datetime import datetime, timezone, timedelta
        now = datetime.now(timezone(timedelta(hours=8)))
        year, month = str(now.year), str(now.month)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT fb.*, fc.name as category_name, fc.type as category_type, fc.color
            FROM finance_budgets fb
            JOIN finance_categories fc ON fc.id=fb.category_id
            WHERE fb.year=%s AND fb.month=%s
            ORDER BY fc.type, fc.sort_order
        """, (int(year), int(month))).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d['budget_amount'] = float(d['budget_amount'])
        if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
        if d.get('updated_at'): d['updated_at'] = d['updated_at'].isoformat()
        result.append(d)
    return jsonify(result)


@bp.route('/api/finance/budgets', methods=['POST'])
@require_module('finance')
def api_budgets_save():
    b     = request.get_json(force=True)
    year  = int(b.get('year',  0))
    month = int(b.get('month', 0))
    items = b.get('items', [])
    if not year or not month: return jsonify({'error': '年月為必填'}), 400
    with get_db() as conn:
        for it in items:
            cid = it.get('category_id')
            amt = float(it.get('budget_amount', 0))
            if cid is None: continue
            if amt == 0:
                conn.execute(
                    "DELETE FROM finance_budgets WHERE year=%s AND month=%s AND category_id=%s",
                    (year, month, cid))
            else:
                conn.execute("""
                    INSERT INTO finance_budgets (year, month, category_id, budget_amount, updated_at)
                    VALUES (%s,%s,%s,%s,NOW())
                    ON CONFLICT (year, month, category_id)
                    DO UPDATE SET budget_amount=EXCLUDED.budget_amount, updated_at=NOW()
                """, (year, month, cid, amt))
    return jsonify({'ok': True})


@bp.route('/api/finance/budgets/vs-actual', methods=['GET'])
@require_module('finance')
def api_budgets_vs_actual():
    year  = request.args.get('year',  '')
    month = request.args.get('month', '')
    if not year or not month:
        from datetime import datetime, timezone, timedelta
        now = datetime.now(timezone(timedelta(hours=8)))
        year, month = str(now.year), str(now.month)
    period = f"{year}-{str(month).zfill(2)}"
    with get_db() as conn:
        cats = conn.execute(
            "SELECT id, name, type, color FROM finance_categories WHERE active=TRUE ORDER BY type, sort_order"
        ).fetchall()
        budgets = conn.execute(
            "SELECT category_id, budget_amount FROM finance_budgets WHERE year=%s AND month=%s",
            (int(year), int(month))
        ).fetchall()
        actuals = conn.execute("""
            SELECT category_id, SUM(amount) as total
            FROM finance_records
            WHERE TO_CHAR(record_date,'YYYY-MM')=%s
            GROUP BY category_id
        """, (period,)).fetchall()
    budget_map = {r['category_id']: float(r['budget_amount']) for r in budgets}
    actual_map = {r['category_id']: float(r['total']) for r in actuals}
    result = []
    for c in cats:
        cid = c['id']
        bgt = budget_map.get(cid, 0)
        act = actual_map.get(cid, 0)
        pct = round(act / bgt * 100, 1) if bgt > 0 else None
        result.append({
            'category_id': cid, 'category_name': c['name'],
            'category_type': c['type'], 'color': c['color'],
            'budget': bgt, 'actual': act,
            'remaining': round(bgt - act, 2),
            'pct': pct, 'over_budget': bgt > 0 and act > bgt,
        })
    return jsonify({'year': year, 'month': month, 'items': result})


# ── Payroll → Finance ──────────────────────────────────────────

@bp.route('/api/finance/payroll/status', methods=['GET'])
@require_module('finance')
def api_payroll_status():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT month,
                   COUNT(*) as total,
                   SUM(CASE WHEN finance_synced THEN 1 ELSE 0 END) as synced,
                   SUM(net_pay) as total_net_pay
            FROM salary_records
            WHERE status IN ('confirmed','draft')
            GROUP BY month ORDER BY month DESC LIMIT 24
        """).fetchall()
    return jsonify([{
        'month':         r['month'],
        'total':         int(r['total']),
        'synced':        int(r['synced']),
        'total_net_pay': float(r['total_net_pay'] or 0),
        'all_synced':    int(r['synced']) == int(r['total']),
    } for r in rows])


@bp.route('/api/finance/payroll/sync', methods=['POST'])
@require_module('finance')
def api_payroll_sync():
    b     = request.get_json(force=True)
    month = b.get('month', '')
    if not month: return jsonify({'error': '請提供月份'}), 400
    with get_db() as conn:
        cat = conn.execute(
            "SELECT id FROM finance_categories WHERE name='薪資支出' AND type='expense' LIMIT 1"
        ).fetchone()
        if not cat:
            cat = conn.execute("""
                INSERT INTO finance_categories (name,type,color,sort_order)
                VALUES ('薪資支出','expense','#e07b2a',11) RETURNING *
            """).fetchone()
        cat_id = cat['id']

        records = conn.execute("""
            SELECT sr.*, ps.name as staff_name
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id=sr.staff_id
            WHERE sr.month=%s AND sr.finance_synced=FALSE
        """, (month,)).fetchall()
        if not records:
            return jsonify({'created': 0, 'message': '無需同步的薪資記錄'})

        import calendar as _cal
        from datetime import date as _d
        first_pay_date = records[0].get('pay_date') if records else None
        if first_pay_date:
            record_date = first_pay_date.isoformat() if hasattr(first_pay_date, 'isoformat') else str(first_pay_date)
        else:
            from blueprints.salary import _get_salary_config
            cfg = _get_salary_config()
            _y, _m = int(month[:4]), int(month[5:])
            _py, _pm = (_y, _m + 1) if _m < 12 else (_y + 1, 1)
            _pd = min(cfg['pay_day'], _cal.monthrange(_py, _pm)[1])
            record_date = _d(_py, _pm, _pd).isoformat()

        created = 0
        for sr in records:
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, note, created_by)
                VALUES (%s,%s,'expense',%s,%s,%s,'payroll-sync')
            """, (record_date, cat_id,
                  f"{sr['staff_name']} {month} 薪資",
                  float(sr['net_pay']),
                  f"薪資記錄 #{sr['id']}"))
            conn.execute("UPDATE salary_records SET finance_synced=TRUE WHERE id=%s", (sr['id'],))
            created += 1
    return jsonify({'created': created, 'month': month})
